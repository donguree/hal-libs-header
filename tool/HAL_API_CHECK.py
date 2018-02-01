import re
import copy
import openpyxl
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border, Side
from openpyxl.styles.colors import WHITE
from openpyxl.worksheet.datavalidation import DataValidation
from paramiko import SSHClient,AutoAddPolicy
import datetime
import sys

FILE_NAME = 'HAL_API_CheckList('+ str(datetime.date.today()) + ').xlsx'

class EXCEL():
    def __init__(self):
        self.wb = openpyxl.Workbook()
        SUMMARY = self.wb.active  # 워크시트 활성화/ 워크북 만들면 디폴트로 무조건 하나는 있다. active하기만 하면됨.
        SUMMARY.title = 'SUMMARY'  # 워크시트 이름 바꾸기
        SUMMARY.sheet_properties.tabColor = '1072BA'  # 탭이름의 배경색 변경 RRGGBB
        self.excelDict = dict()
        self.BOX = Border(left=Side(border_style='thin', color='000000'),
                     right=Side(border_style='thin', color='000000'),
                     top=Side(border_style='thin', color='000000'),
                     bottom=Side(border_style='thin', color='000000')
                     )
    def __del__(self):
        self.wb.save(FILE_NAME)
        self.wb.close()

    def summary(self):
        SUMMARY = self.wb['SUMMARY']
        SUMMARY['a1'] = 'Summary'
        SUMMARY['a2'] = 'Created on ' + str(datetime.date.today())
        SUMMARY['a3'] = 'Module'
        SUMMARY['b3'] = 'Owner'
        SUMMARY['c3'] = 'API Total'
        SUMMARY['d3'] = 'API Supported'
        SUMMARY['E3'] = 'OK'
        SUMMARY['F3'] = 'NG'
        SUMMARY['G3'] = 'NT'
        SUMMARY['H3'] = 'NA'
        SUMMARY['J3'] = 'Header Total'
        SUMMARY['k3'] = 'API Total'
        SUMMARY['L3'] = 'API Supported'
        SUMMARY['M3'] = 'Pass Rate(%)'
        SUMMARY['J4'] = len(self.wb.sheetnames)-1
        SUMMARY['k4'] = '=sum(c4:c' + str(len(self.wb.sheetnames)) + ')'
        SUMMARY['l4'] = '=sum(d4:d' + str(len(self.wb.sheetnames)) + ')'
        SUMMARY['m4'] = '=sum(e4:e' + str(len(self.wb.sheetnames)) + ')/L4'

        for idx in range(1,len(self.wb.sheetnames)):
            ws = self.wb[self.wb.sheetnames[idx]]
            if ws.title == 'SUMMARY':
                continue
            SUMMARY['A' + str(idx + 3)] = ws.title
            SUMMARY['c' + str(idx + 3)] = '=' + ws.title + '!D2'
            SUMMARY['d' + str(idx + 3)] = '=' + ws.title + '!E2'
            SUMMARY['e' + str(idx + 3)] = '=' + ws.title + '!F2'
            SUMMARY['f' + str(idx + 3)] = '=' + ws.title + '!G2'
            SUMMARY['g' + str(idx + 3)] = '=' + ws.title + '!H2'
            SUMMARY['H' + str(idx + 3)] = '=' + ws.title + '!I2'

    def decorateExcel(self):
        BOX_GREEN = Border(left=Side(border_style='thin', color='98bd47'),
                          right=Side(border_style='thin', color='98bd47'),
                            top=Side(border_style='thin', color='98bd47'),
                         bottom=Side(border_style='thin', color='98bd47')
                          )
        BOX_BLUE = Border(left=Side(border_style='thin', color='5083c0'),
                         right=Side(border_style='thin', color='5083c0'),
                           top=Side(border_style='thin', color='5083c0'),
                        bottom=Side(border_style='thin', color='5083c0')
                          )
        BOX_RED = Border(left=Side(border_style='thin', color='C03737'),
                        right=Side(border_style='thin', color='C03737'),
                          top=Side(border_style='thin', color='C03737'),
                       bottom=Side(border_style='thin', color='C03737')
                          )
        for ws in self.wb:
            if ws.title == 'SUMMARY':
                # 셀 너비
                ws.column_dimensions["a"].width = 20
                for i in ('b', 'c', 'd'):
                    ws.column_dimensions[i].width = 14
                for i in ('e', 'f', 'g', 'h'):
                    ws.column_dimensions[i].width = 8
                for i in ('j', 'k', 'l', 'm'):
                    ws.column_dimensions[i].width = 13.5
                # 셀 폰트
                ws['a1'].font = Font(b=True, size=14)
                for i in ('a','b','c','d','e','f','g','h'):
                    ws[i+str(3)].font = Font(b=True, color=WHITE)
                    ws[i+str(3)].fill = PatternFill(fgColor=Color('5083c0'), patternType='solid')
                for i in ('j', 'k', 'l', 'm'):
                    ws[i+str(3)].font = Font(b=True, color=WHITE)
                    ws[i+str(3)].fill = PatternFill(fgColor=Color('C03737'), patternType='solid')
                # 셀 얼라인
                for i in ('a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i','j','k','l','m'):
                    ws[i + str(3)].alignment = Alignment(horizontal='center', vertical='center')
                # 셀 고정
                ws.freeze_panes = 'A4'
                # 테두리
                for j in range(4,ws.max_row+1):
                    for i in ('a', 'b', 'c','d','e','f','g','h'):
                        ws[i+str(j)].border = BOX_BLUE
                for i in ('j', 'k', 'l', 'm'):
                    ws[i + str(4)].border = BOX_RED

            else:
                # 셀 너비
                ws.column_dimensions["a"].width = 50
                ws.column_dimensions["b"].width = 60
                ws.column_dimensions["c"].width = 15
                for i in ('d', 'e', 'f', 'g', 'h', 'i'):
                    ws.column_dimensions[i].width = 7

                # 셀 폰트
                ws['a1'].font = Font(b=True, size=14)
                for i in ('a','b','c'):
                    ws[i+str(2)].font = Font(b=True, color=WHITE)
                    ws[i+str(2)].fill = PatternFill(fgColor=Color('98bd47'), patternType='solid')
                for i in ('d','e','f','g','h','i'):
                    ws[i+str(1)].font = Font(b=True, color=WHITE)
                    ws[i+str(1)].fill = PatternFill(fgColor=Color('5083c0'), patternType='solid')

                # 셀 얼라인
                for i in ('a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i'):
                    ws[i+str(2)].alignment= Alignment(horizontal='center', vertical='center')
                for i in ('d', 'e', 'f', 'g', 'h', 'i'):
                    ws[i + str(1)].alignment = Alignment(horizontal='center', vertical='center')

                # 셀 고정
                ws.freeze_panes = 'A3'

                # 테두리
                for j in range(3,ws.max_row+1):
                    for i in ('a', 'b', 'c'):
                        ws[i+str(j)].border = BOX_GREEN
                for i in ('d', 'e', 'f', 'g', 'h', 'i'):
                    ws[i + str(2)].border = BOX_BLUE

                # 필터
                ws.auto_filter.ref = "A2:c"+str(ws.max_row+1)

                # 값 고정 (data-validation)
                dv = DataValidation(type='list', formula1='"OK,NG,NA,NT"', allow_blank=True)
                dv.error = 'OK, NG, NA, NT 만 기입할 수 있습니다. / Your entry is not in the list(OK, NG, NA, NT)'
                dv.errorTitle = 'Invalid Entry'
                dv.prompt = 'Please select from the list (OK, NG, NT(need to support but not yet ready), NA(not support)'
                dv.promptTitle = 'List Selection'
                ws.add_data_validation(dv)
                dv.ranges.append('c3:c' + str(ws.max_row ))

    def writeData_form_GIT(self, gitData):
        for module, api in gitData.items():
            ws = self.wb.create_sheet(module)
            for index in range(len(api)):
                ws['a1'] = module
                ws['a2'] = 'HAL API'
                ws['b2'] = 'Test Method'
                ws['c2'] = 'Test Result'
                ws['d1'] = 'Total'
                ws['e1'] = 'Support'
                ws['f1'] = 'OK'
                ws['g1'] = 'NG'
                ws['h1'] = 'NT'
                ws['I1'] = 'NA'
                ws['a'+str(index+3)] = api[index]
                ws['d2'] = '=SUM(F2:I2)'
                ws['e2'] = '=D2-I2'
                ws['f2'] = '=COUNTIF(C3:C' + str(ws.max_row) + ',"OK")'
                ws['g2'] = '=COUNTIF(C3:C' + str(ws.max_row) + ',"NG")'
                ws['H2'] = '=COUNTIF(C3:C' + str(ws.max_row) + ',"NT")'
                ws['i2'] = '=COUNTIF(C3:C' + str(ws.max_row) + ',"NA")'

class GIT():

    def __init__(self,ID=None, PW=None, URL=None):
        # 로그인 정보
        self.ID  = ID
        self.PW  = PW
        self.url = URL
        self.basic_path = '//'+ self.url + '/work/hal_header/hal-libs-header/hal_inc/'
        self.gitData = dict()

    def sshConnect(self):
        client = SSHClient()
        client.load_system_host_keys()
        client.set_missing_host_key_policy(AutoAddPolicy())
        client.connect(self.url, username=self.ID, password=self.PW)
        cmd_list = ['cd hal_header/hal-libs-header/hal_inc/', 'ls']
        cmd_list = ';'.join(cmd_list)
        #
        stdin, stdout, stderr = client.exec_command(cmd_list)
        return stdout

    def findHALLData(self):
        stdout = self.sshConnect()
        #헤더파일 하나씩 읽기
        for header in stdout:
            header = header.strip()
            file_path = self.basic_path + header
            try:
                fp = open(file_path, 'rt')
                targetLine = re.findall('[a-zA-Z*_\d]+[*]*[" "\t]+[*]*HAL_[a-zA-Z_]+[()*\s\w/,]+;', fp.read())
            except Exception as e:
                fp = open(file_path, 'rt', encoding='UTF8')
                targetLine = re.findall('[a-zA-Z*_\d]+[*]*[" "\t]+[*]*HAL_[a-zA-Z_]+[()*\s\w/,]+;', fp.read())

            for i in targetLine:
                # 1. 함수명이 있는 line 찾기
                pattern = re.compile(r'\s+')
                target = re.sub(pattern, " ", i).lstrip()  # bingo
                if '(' not in target:
                    continue
                splitedTarget = re.split('\(', target)

                # 2. Func Name 찾기
                # extern 제거
                if 'extern' in splitedTarget[0]:
                    splitedTarget[0] = splitedTarget[0].split('extern ')[1]

                if '*' in splitedTarget[0]:
                    funcName = re.split("\*", splitedTarget[0])[1]
                else:
                    tmp = re.split(" ", splitedTarget[0])
                    funcName = tmp[1]
                if header not in self.gitData.keys():
                    self.gitData[header] = list()
                self.gitData[header].append(funcName)
                self.gitData[header] = sorted(self.gitData[header])

        return copy.deepcopy(self.gitData)


def main():
    if len(sys.argv) is 1:
        print("ID(OE server) PW 입력해주세요.(ID PW)")
        git = GIT()
        print('1/6 GIT object was created.')
    else:
        git = GIT(sys.argv[1], sys.argv[2], sys.argv[3])
        print('1/6 GIT object was created.')
    gitData = git.findHALLData()
    print('2/6 Found HAL API from git')
    excel = EXCEL()
    print('3/6 EXCEL object was created.')
    excel.writeData_form_GIT(gitData)
    print('4/6 Wrote data to a Excel Doc.')
    excel.decorateExcel()
    print('5/6 Decorated Excel style')
    excel.summary()
    print('6/6 Made a summary in the Excel')


if __name__ == '__main__':
    main()